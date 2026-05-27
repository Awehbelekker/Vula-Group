"""
vula/takeoff/boq_export.py

Generates a formatted Excel workbook from a BOQ dict.
Returns a BytesIO so the API can stream it directly.

Sheets:
    1. Summary  — project info, totals, markup
    2. Full BOQ — all items grouped by trade
    3. By Trade — one row per trade with subtotals
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Palette ────────────────────────────────────────────────────────────────────
GREEN      = "FF2C5545"
GREEN_LIGHT = "FFE8F0EC"
AMBER      = "FFD97706"
AMBER_LIGHT = "FFFFF8E7"
GREY       = "FFF0EDE5"
WHITE      = "FFFFFFFF"
BLACK      = "FF1E1E1E"
RED_LIGHT  = "FFFFF0F0"


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10, color: str = BLACK, name: str = "Calibri") -> "Font":
    return Font(bold=bold, size=size, color=color, name=name)


def _border() -> "Border":
    thin = Side(style="thin", color="FFCCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _currency(ws: Any, cell: Any, value: float) -> None:
    cell.value = value
    cell.number_format = 'R #,##0.00'
    cell.alignment = Alignment(horizontal="right")


def generate_boq_excel(boq: dict, project_name: str = "") -> io.BytesIO:
    """
    Build a formatted Excel workbook from a BOQ dict (as returned by BOQ.to_dict()).
    Returns a BytesIO containing the .xlsx file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _build_summary_sheet(wb, boq, project_name)
    _build_boq_sheet(wb, boq)
    _build_trade_summary_sheet(wb, boq)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_summary_sheet(wb: Any, boq: dict, project_name: str) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "BILL OF QUANTITIES"
    c.font = _font(bold=True, size=18, color="FFFFFFFF")
    c.fill = _fill(GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-title
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = project_name or boq.get("project_name", "Vula Project")
    c.font = _font(bold=True, size=13)
    c.fill = _fill(GREEN_LIGHT)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Generated date
    ws.merge_cells("A3:B3")
    c = ws["A3"]
    c.value = f"Generated: {datetime.now().strftime('%d %B %Y')}"
    c.font = _font(size=9, color="FF8A8680")
    c.alignment = Alignment(horizontal="center")

    rows = [
        ("", ""),
        ("Net Total (excl. markup)", boq.get("net_total", 0)),
        (f"Markup ({boq.get('markup_pct', 15):.0f}%)", boq.get("markup_amount", 0)),
        ("Grand Total", boq.get("grand_total", 0)),
        ("", ""),
        ("VAT (15%)", boq.get("grand_total", 0) * 0.15),
        ("Total incl. VAT", boq.get("grand_total", 0) * 1.15),
    ]

    for i, (label, value) in enumerate(rows, start=5):
        row = ws.row_dimensions[i]
        row.height = 20
        la = ws.cell(row=i, column=1, value=label)
        la.font = _font(bold=(label in ("Grand Total", "Total incl. VAT")), size=11)
        la.border = _border()
        if isinstance(value, (int, float)) and value:
            va = ws.cell(row=i, column=2)
            _currency(ws, va, value)
            va.font = _font(bold=(label in ("Grand Total", "Total incl. VAT")), size=11)
            va.border = _border()
            if label in ("Grand Total", "Total incl. VAT"):
                la.fill = _fill(GREEN_LIGHT)
                va.fill = _fill(GREEN_LIGHT)

    # Trade breakdown
    by_trade = boq.get("by_trade", {})
    if by_trade:
        r = len(rows) + 6
        ws.cell(row=r, column=1, value="TRADE BREAKDOWN").font = _font(bold=True, size=11, color="FFFFFFFF")
        ws.cell(row=r, column=1).fill = _fill(GREEN)
        ws.cell(row=r, column=2).fill = _fill(GREEN)
        r += 1
        for trade, items in by_trade.items():
            total = sum(it.get("line_total", 0) for it in items)
            la = ws.cell(row=r, column=1, value=trade)
            la.font = _font(size=10)
            la.border = _border()
            va = ws.cell(row=r, column=2)
            _currency(ws, va, total)
            va.border = _border()
            r += 1


def _build_boq_sheet(wb: Any, boq: dict) -> None:
    ws = wb.create_sheet("Full BOQ")

    headers = ["#", "Trade", "Description", "Unit", "Qty", "Rate (R)", "Line Total (R)", "Source", "Confidence"]
    widths   = [4,   18,     40,             8,      10,    14,         16,               14,        12]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = _font(bold=True, size=10, color="FFFFFFFF")
        c.fill = _fill(GREEN)
        c.alignment = Alignment(horizontal="center")
        c.border = _border()
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    items = boq.get("items", [])
    for i, item in enumerate(items, start=1):
        row = i + 1
        fill = _fill(GREY) if i % 2 == 0 else _fill(WHITE)
        values = [
            i,
            item.get("trade", ""),
            item.get("description", ""),
            item.get("unit", ""),
            item.get("qty", 0),
            item.get("rate_used", 0),
            item.get("line_total", 0),
            item.get("source", ""),
            f"{item.get('confidence', 0):.0f}%",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.border = _border()
            c.font = _font(size=9)
            if col in (5, 6, 7):
                c.alignment = Alignment(horizontal="right")
            if col in (6, 7):
                c.number_format = 'R #,##0.00'

    # Totals row
    total_row = len(items) + 2
    ws.cell(row=total_row, column=3, value="GRAND TOTAL").font = _font(bold=True, size=10)
    ws.cell(row=total_row, column=3).fill = _fill(GREEN_LIGHT)
    t = ws.cell(row=total_row, column=7, value=boq.get("grand_total", 0))
    t.number_format = 'R #,##0.00'
    t.font = _font(bold=True, size=10)
    t.fill = _fill(GREEN_LIGHT)
    t.alignment = Alignment(horizontal="right")


def _build_trade_summary_sheet(wb: Any, boq: dict) -> None:
    ws = wb.create_sheet("Trade Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    headers = ["Trade", "Items", "Subtotal (R)", "% of Net"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _font(bold=True, size=10, color="FFFFFFFF")
        c.fill = _fill(GREEN)
        c.border = _border()
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    by_trade = boq.get("by_trade", {})
    net_total = boq.get("net_total", 1) or 1
    for row, (trade, items) in enumerate(by_trade.items(), start=2):
        subtotal = sum(it.get("line_total", 0) for it in items)
        fill = _fill(GREY) if row % 2 == 0 else _fill(WHITE)
        ws.cell(row=row, column=1, value=trade).font = _font(size=10)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = _border()
        ws.cell(row=row, column=2, value=len(items)).fill = fill
        ws.cell(row=row, column=2).border = _border()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        c = ws.cell(row=row, column=3, value=subtotal)
        c.number_format = 'R #,##0.00'
        c.alignment = Alignment(horizontal="right")
        c.fill = fill
        c.border = _border()
        pct = ws.cell(row=row, column=4, value=subtotal / net_total)
        pct.number_format = '0.0%'
        pct.alignment = Alignment(horizontal="center")
        pct.fill = fill
        pct.border = _border()

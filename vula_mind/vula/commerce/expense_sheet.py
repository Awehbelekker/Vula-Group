"""
vula/commerce/expense_sheet.py — per-rep monthly expense claim sheet.

Compiles a rep's claimed expenses for the prior calendar month into a formatted .xlsx matching
the REAL claim sheet Ian already uses by hand (a shared .xlsx, 2026-08-26): a "Recon" summary
sheet grouping claims into PETROL / CLIENTS (client refreshments) / ACCOMMODATION — plus an
OTHER catch-all for anything that doesn't fit — with a per-category sheet holding the actual
slip photos. Emailed on the rep's own configured day of the month (migration 140).

Category classification: commerce_expenses.purpose_category, auto-classified at scan time
(vula.commerce.expenses.classify_purpose_category — deterministic vendor match first, one LLM
call only when that's inconclusive) or resolved from a rep's WhatsApp reply when genuinely
uncertain. Falls back to the chart-of-accounts account_code (fuel → petrol) for any older claim
that predates purpose_category ever being set.

Petrol claims also carry odometer_km (migration 142) — a real KM logbook column from Ian's own
original claim-sheet template, asked for automatically once a receipt classifies as petrol.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from vula.commerce import expenses, service

log = logging.getLogger(__name__)

PURPOSE_CATEGORIES = ["petrol", "clients", "accommodation", "other"]
_CATEGORY_ORDER = ["petrol", "clients", "accommodation", "other"]
_CATEGORY_LABEL = {"petrol": "PETROL", "clients": "CLIENTS", "accommodation": "ACCOMMODATION", "other": "OTHER"}
_CATEGORY_SHEET_TITLE = {"petrol": "Petrol slips", "clients": "Clients slips",
                        "accommodation": "Accommodation slips", "other": "Other slips"}

GREEN = "FF2C5545"
GREEN_LIGHT = "FFE8F0EC"
GREY = "FFF0EDE5"
WHITE = "FFFFFFFF"
BLACK = "FF1E1E1E"


def _client():
    return service._client()


def _infer_purpose_category(claim: Dict[str, Any]) -> str:
    """purpose_category if a rep has already been asked (not yet built); otherwise a
    deterministic guess from the account_code the existing scan pipeline already assigns —
    never an LLM guess, matching this platform's "constrained classification, real data only"
    discipline everywhere else."""
    explicit = (claim.get("purpose_category") or "").strip().lower()
    if explicit in PURPOSE_CATEGORIES:
        return explicit
    if (claim.get("account_code") or "").strip().lower() == "fuel":
        return "petrol"
    return "other"


def prior_month_range(now: Optional[datetime] = None) -> tuple[str, str]:
    """(start, end) ISO dates for the full calendar month before `now` — the month a claim
    sheet fired on day_of_month is meant to summarize (day 1 of THIS month has no expenses in
    it yet; the sheet always covers the month that just ended, regardless of which day the rep
    chose to receive it on)."""
    now = now or datetime.now(timezone.utc)
    first_of_this_month = date(now.year, now.month, 1)
    last_of_prior_month = first_of_this_month - timedelta(days=1)
    first_of_prior_month = date(last_of_prior_month.year, last_of_prior_month.month, 1)
    return first_of_prior_month.isoformat(), last_of_prior_month.isoformat()


async def fetch_claims(tenant_id: str, paid_by: str, month_start: str, month_end: str) -> List[Dict[str, Any]]:
    try:
        rows = (_client().table("commerce_expenses").select("*")
                .eq("tenant_id", tenant_id).eq("paid_by", paid_by)
                .gte("date", month_start).lte("date", month_end)
                .order("date", desc=False).execute().data or [])
        return rows
    except Exception as exc:
        log.warning("expense sheet claim fetch failed for %s/%s: %s", tenant_id, paid_by, exc)
        return []


async def _fetch_image_bytes(url: str) -> Optional[bytes]:
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(url)
            if resp.is_success:
                return resp.content
    except Exception as exc:
        log.debug("receipt image fetch failed for %s: %s", url, exc)
    return None


async def build_expense_workbook(tenant_id: str, rep_name: str, paid_by: str,
                                 month_start: str, month_end: str) -> Optional[bytes]:
    """The Recon-shaped workbook. Returns None if the rep has no claims that month (a caller
    should skip sending rather than mail an empty sheet)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        log.warning("openpyxl not installed — cannot build expense workbook")
        return None

    claims = await fetch_claims(tenant_id, paid_by, month_start, month_end)
    if not claims:
        return None

    by_category: Dict[str, List[Dict[str, Any]]] = {c: [] for c in _CATEGORY_ORDER}
    for c in claims:
        by_category[_infer_purpose_category(c)].append(c)

    wb = openpyxl.Workbook()
    recon = wb.active
    recon.title = "Recon"

    def fill(hex_color: str) -> "PatternFill":
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, size=10, color=BLACK) -> "Font":
        return Font(bold=bold, size=size, color=color, name="Calibri")

    period_label = f"{month_start} to {month_end}"
    recon["B2"] = f"ADVANCE CLAIMS - {rep_name.upper()}"
    recon["B2"].font = font(bold=True, size=14, color=GREEN)
    recon["B3"] = period_label
    recon["B3"].font = font(size=10)

    row = 5
    grand_total_cents = 0
    for cat in _CATEGORY_ORDER:
        items = by_category[cat]
        if not items:
            continue
        is_petrol = cat == "petrol"
        recon.cell(row=row, column=2, value=_CATEGORY_LABEL[cat]).font = font(bold=True, size=11, color=WHITE)
        recon.cell(row=row, column=2).fill = fill(GREEN)
        row += 1
        # KM/KM since last — real columns from Ian's own original claim-sheet template, only
        # meaningful for petrol (a South African travel-logbook expectation, migration 142).
        headers = ["Date", "Vendor", "Amount", "KM", "KM since last"] if is_petrol else ["Date", "Vendor", "Amount"]
        for i, h in enumerate(headers):
            cell = recon.cell(row=row, column=2 + i, value=h)
            cell.font = font(bold=True, size=9)
            cell.fill = fill(GREY)
        row += 1
        cat_total = 0
        prev_km: Optional[int] = None
        if is_petrol and items:
            prev_km = expenses.last_odometer_before(
                tenant_id, paid_by, items[0]["date"], exclude_id=items[0].get("id"))
        for c in items:
            amt = int(c.get("amount_cents") or 0)
            cat_total += amt
            recon.cell(row=row, column=2, value=c.get("date"))
            recon.cell(row=row, column=3, value=c.get("supplier") or "—")
            amt_cell = recon.cell(row=row, column=4, value=amt / 100)
            amt_cell.number_format = 'R #,##0.00'
            amt_cell.alignment = Alignment(horizontal="right")
            if is_petrol:
                km = c.get("odometer_km")
                recon.cell(row=row, column=5, value=km if km is not None else "—")
                delta = (km - prev_km) if (km is not None and prev_km is not None and km > prev_km) else None
                recon.cell(row=row, column=6, value=delta if delta is not None else "—")
                if km is not None:
                    prev_km = km
            row += 1
        total_cell = recon.cell(row=row, column=3, value=f"{_CATEGORY_LABEL[cat]} TOTAL")
        total_cell.font = font(bold=True, size=9)
        amt_cell = recon.cell(row=row, column=4, value=cat_total / 100)
        amt_cell.number_format = 'R #,##0.00'
        amt_cell.font = font(bold=True)
        amt_cell.alignment = Alignment(horizontal="right")
        grand_total_cents += cat_total
        row += 2

    recon.cell(row=row, column=3, value="GRAND TOTAL").font = font(bold=True, size=11)
    gt_cell = recon.cell(row=row, column=4, value=grand_total_cents / 100)
    gt_cell.number_format = 'R #,##0.00'
    gt_cell.font = font(bold=True, size=11, color=GREEN)
    gt_cell.alignment = Alignment(horizontal="right")

    for col, width in (("A", 3), ("B", 16), ("C", 28), ("D", 14), ("E", 12), ("F", 14)):
        recon.column_dimensions[col].width = width

    # One sheet per non-empty category, with the real slip photos embedded — matching Ian's own
    # existing manual workbook shape ("Petrol slips" / "Clients slips" / "Accom inv") exactly,
    # rather than a generic single "attachments" list.
    for cat in _CATEGORY_ORDER:
        items = by_category[cat]
        if not items:
            continue
        ws = wb.create_sheet(_CATEGORY_SHEET_TITLE[cat])
        img_row = 1
        for c in items:
            label = f"{c.get('date')} — {c.get('supplier') or 'Unknown'} — R{int(c.get('amount_cents') or 0) / 100:.2f}"
            ws.cell(row=img_row, column=1, value=label).font = font(bold=True, size=10)
            img_row += 1
            url = c.get("receipt_url")
            if url:
                data = await _fetch_image_bytes(url)
                if data:
                    try:
                        img = XLImage(io.BytesIO(data))
                        img.width, img.height = 240, 180
                        ws.add_image(img, f"A{img_row}")
                        img_row += 12  # rows-worth of vertical space for the embedded image
                    except Exception as exc:
                        log.debug("could not embed receipt image for %s: %s", c.get("id"), exc)
            img_row += 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def send_expense_sheet(tenant_id: str, rep: Dict[str, Any]) -> bool:
    """Compile + email the rep's prior-month claim sheet. Never raises — a missing SMTP
    connection or a genuinely empty month both degrade to a clean False, matching
    vula/commerce/call_sheet.py's established send-path discipline."""
    recipient = rep.get("expense_sheet_recipient_email")
    if not recipient:
        return False
    month_start, month_end = prior_month_range()
    rep_name = rep.get("name") or "Rep"
    workbook = await build_expense_workbook(tenant_id, rep_name, rep["whatsapp"], month_start, month_end)
    if not workbook:
        log.info("expense sheet skipped for %s/%s — no claims for %s to %s",
                 tenant_id, rep["whatsapp"], month_start, month_end)
        return False

    subject = f"Expense Claim — {rep_name} — {month_start[:7]}"
    body = (f"Attached is {rep_name}'s expense claim sheet for {month_start} to {month_end}, "
            "compiled automatically from receipts scanned via WhatsApp.")
    from vula.commerce.mail_router import send_tenant_email
    sent = await send_tenant_email(tenant_id, recipient, subject, body, attachments=[
        {"filename": f"expense-claim-{month_start[:7]}.xlsx", "content": workbook,
         "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    ])
    if not sent:
        log.info("expense sheet email not sent for %s/%s — no working connected mailbox "
                 "(IMAP or Microsoft)", tenant_id, rep["whatsapp"])
    return sent


def is_due(rep: Dict[str, Any], now: datetime) -> bool:
    """Fires once, on the rep's configured day of month, not more than once per calendar
    month — mirrors call_sheet.py's is_due() shape, keyed on day-of-month instead of
    day-of-week."""
    target_day = rep.get("expense_sheet_day_of_month") or 1
    if now.day != target_day:
        return False
    last = rep.get("expense_sheet_last_sent_at")
    if last:
        try:
            last_dt = datetime.fromisoformat(str(last).replace("Z", "+00:00"))
            if last_dt.year == now.year and last_dt.month == now.month:
                return False
        except Exception:
            pass
    return True


async def run_monthly_expense_sheets() -> int:
    """For every sales_rep with a monthly-sheet recipient configured and due today: compile and
    send. Never lets one rep's failure stop another — same per-row defensive discipline as
    call_sheet.py's run_weekly_call_sheets()."""
    try:
        reps = (_client().table("vula_team_members").select(
                    "id,tenant_id,whatsapp,name,expense_sheet_recipient_email,"
                    "expense_sheet_day_of_month,expense_sheet_last_sent_at")
                .eq("role", "sales_rep").eq("active", True)
                .execute().data or [])
    except Exception as exc:
        log.warning("expense sheet rep query failed: %s", exc)
        return 0

    now = datetime.now(timezone.utc)
    sent_count = 0
    for rep in reps:
        if not rep.get("expense_sheet_recipient_email"):
            continue
        try:
            if not is_due(rep, now):
                continue
            ok = await send_expense_sheet(rep["tenant_id"], rep)
            # A due-and-handled cycle (even a genuinely empty month) counts as "handled for this
            # cycle" — the configured day-of-month is a calendar slot, not a rolling window, so
            # stamping regardless of outcome is what keeps a rep who logged nothing this month
            # from being re-sent/re-checked, matching call_sheet.py's run_weekly_call_sheets().
            _client().table("vula_team_members").update(
                {"expense_sheet_last_sent_at": now.isoformat()}
            ).eq("id", rep["id"]).execute()
            if ok:
                sent_count += 1
        except Exception as exc:
            log.warning("expense sheet failed for rep %s/%s: %s", rep.get("tenant_id"), rep.get("whatsapp"), exc)
    return sent_count

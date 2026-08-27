"""Tests for vula/commerce/expense_sheet.py — the per-rep monthly expense claim workbook
(migration 140). Uses the same in-memory fake Supabase client pattern as
tests/test_call_sheet.py.
"""
import uuid
from datetime import datetime, timedelta, timezone
from unittest.mock import AsyncMock, patch

import pytest

import vula.commerce.expense_sheet as es_mod
from vula.commerce import expenses

TID = "test-tenant"
REP_PHONE = "27821234567"


class _Result:
    def __init__(self, data):
        self.data = data


class _NotProxy:
    def __init__(self, query):
        self._query = query

    def is_(self, key, _val):
        self._query._not_null.append(key)
        return self._query


class _FakeQuery:
    def __init__(self, table):
        self.table = table
        self.filters = []
        self._gte = []
        self._lte = []
        self._lt = []
        self._not_null = []
        self._limit = None
        self._patch = None

    def select(self, *_a, **_kw):
        return self

    def eq(self, key, val):
        self.filters.append((key, val))
        return self

    def gte(self, key, val):
        self._gte.append((key, val))
        return self

    def lte(self, key, val):
        self._lte.append((key, val))
        return self

    def lt(self, key, val):
        self._lt.append((key, val))
        return self

    @property
    def not_(self):
        return _NotProxy(self)

    def order(self, *_a, **_kw):
        return self

    def limit(self, n):
        self._limit = n
        return self

    def _matches(self, row):
        if not all(row.get(k) == v for k, v in self.filters):
            return False
        if not all((row.get(k) or "") >= v for k, v in self._gte):
            return False
        if not all((row.get(k) or "") <= v for k, v in self._lte):
            return False
        if not all((row.get(k) or "") < v for k, v in self._lt):
            return False
        if not all(row.get(k) is not None for k in self._not_null):
            return False
        return True

    def insert(self, row):
        row = dict(row)
        row.setdefault("id", str(uuid.uuid4()))
        self.table.rows.append(row)
        return _ExecWrapper([row])

    def update(self, patch_dict):
        self._patch = patch_dict
        return self

    def execute(self):
        rows = [r for r in self.table.rows if self._matches(r)]
        if self._patch is not None:
            for r in rows:
                r.update(self._patch)
        if self._limit:
            rows = rows[: self._limit]
        return _Result(rows)


class _ExecWrapper:
    def __init__(self, rows):
        self._rows = rows

    def execute(self):
        return _Result(self._rows)


class _FakeTable:
    def __init__(self, name, store):
        self.rows = store.setdefault(name, [])


class _FakeClient:
    def __init__(self):
        self.store = {}

    def table(self, name):
        return _FakeQuery(_FakeTable(name, self.store))


@pytest.fixture
def fake_client(monkeypatch):
    client = _FakeClient()
    monkeypatch.setattr(es_mod, "_client", lambda: client)
    # last_odometer_before() is called directly on the expenses module (not through
    # expense_sheet.py's own _client wrapper) — same in-memory store, or it'd silently hit
    # the real DB from a unit test.
    monkeypatch.setattr(expenses, "_client", lambda: client)
    return client


def _rep(**over):
    row = {"id": str(uuid.uuid4()), "tenant_id": TID, "whatsapp": REP_PHONE, "name": "Richard",
           "role": "sales_rep", "active": True,
           "expense_sheet_recipient_email": "accounts@gerflor.co.za",
           "expense_sheet_day_of_month": 1, "expense_sheet_last_sent_at": None}
    row.update(over)
    return row


def _claim(**over):
    row = {"id": str(uuid.uuid4()), "tenant_id": TID, "paid_by": REP_PHONE,
           "date": "2026-07-15", "supplier": "Engen", "amount_cents": 55000,
           "account_code": "fuel", "purpose_category": None, "receipt_url": None}
    row.update(over)
    return row


# ── prior_month_range ──────────────────────────────────────────────────────────────

def test_prior_month_range_returns_full_prior_calendar_month():
    now = datetime(2026, 8, 15, tzinfo=timezone.utc)
    start, end = es_mod.prior_month_range(now)
    assert start == "2026-07-01"
    assert end == "2026-07-31"


def test_prior_month_range_handles_january_rollover():
    now = datetime(2026, 1, 10, tzinfo=timezone.utc)
    start, end = es_mod.prior_month_range(now)
    assert start == "2025-12-01"
    assert end == "2025-12-31"


# ── _infer_purpose_category ──────────────────────────────────────────────────────────

def test_infer_purpose_category_uses_explicit_value():
    assert es_mod._infer_purpose_category({"purpose_category": "clients"}) == "clients"


def test_infer_purpose_category_falls_back_to_fuel_account_code():
    assert es_mod._infer_purpose_category({"account_code": "fuel"}) == "petrol"


def test_infer_purpose_category_defaults_to_other():
    assert es_mod._infer_purpose_category({"account_code": "office_supplies"}) == "other"


def test_infer_purpose_category_ignores_invalid_explicit_value():
    assert es_mod._infer_purpose_category({"purpose_category": "bogus", "account_code": "fuel"}) == "petrol"


# ── fetch_claims ──────────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_fetch_claims_scopes_to_tenant_rep_and_month(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-07-10"),
        _claim(date="2026-06-30"),  # outside range
        _claim(date="2026-07-20", paid_by="27829999999"),  # different rep
        _claim(date="2026-07-20", tenant_id="other-tenant"),  # different tenant
    ]
    claims = await es_mod.fetch_claims(TID, REP_PHONE, "2026-07-01", "2026-07-31")
    assert len(claims) == 1
    assert claims[0]["date"] == "2026-07-10"


@pytest.mark.asyncio
async def test_fetch_claims_returns_empty_on_db_error(fake_client, monkeypatch):
    def _boom(*_a, **_kw):
        raise RuntimeError("db down")
    monkeypatch.setattr(fake_client, "table", _boom)
    claims = await es_mod.fetch_claims(TID, REP_PHONE, "2026-07-01", "2026-07-31")
    assert claims == []


# ── build_expense_workbook ────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_build_expense_workbook_returns_none_when_no_claims(fake_client):
    result = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    assert result is None


@pytest.mark.asyncio
async def test_build_expense_workbook_builds_real_xlsx_bytes(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-07-10", supplier="Engen", amount_cents=55000, account_code="fuel"),
        _claim(date="2026-07-12", supplier="Woolworths", amount_cents=12000, purpose_category="clients"),
    ]
    data = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    assert data is not None
    assert isinstance(data, bytes)

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Recon" in wb.sheetnames
    assert "Petrol slips" in wb.sheetnames
    assert "Clients slips" in wb.sheetnames
    assert "Accommodation slips" not in wb.sheetnames  # no accommodation claims this month


@pytest.mark.asyncio
async def test_build_expense_workbook_petrol_table_has_km_columns(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-07-10", supplier="Engen", amount_cents=55000,
               purpose_category="petrol", odometer_km=45280, id="c1"),
        _claim(date="2026-07-20", supplier="Shell", amount_cents=60000,
               purpose_category="petrol", odometer_km=45930, id="c2"),
    ]
    import io
    import openpyxl
    data = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    recon = wb["Recon"]
    values = [[c.value for c in row] for row in recon.iter_rows()]
    flat = [str(v) for row in values for v in row if v is not None]
    assert "KM" in flat
    assert "KM since last" in flat
    assert 45280 in [v for row in values for v in row]
    assert 45930 in [v for row in values for v in row]
    assert 650 in [v for row in values for v in row]  # 45930 - 45280 delta


@pytest.mark.asyncio
async def test_build_expense_workbook_non_petrol_table_has_no_km_columns(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-07-10", supplier="Woolworths", amount_cents=12000, purpose_category="clients"),
    ]
    import io
    import openpyxl
    data = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    recon = wb["Recon"]
    values = [str(c.value) for row in recon.iter_rows() for c in row if c.value is not None]
    assert "KM" not in values


@pytest.mark.asyncio
async def test_build_expense_workbook_petrol_seeds_delta_from_prior_month(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-06-25", supplier="Engen", amount_cents=50000,
               purpose_category="petrol", odometer_km=44000, id="c0"),
        _claim(date="2026-07-05", supplier="Shell", amount_cents=55000,
               purpose_category="petrol", odometer_km=44500, id="c1"),
    ]
    import io
    import openpyxl
    data = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    recon = wb["Recon"]
    values = [v for row in recon.iter_rows() for v in [c.value for c in row]]
    assert 500 in values  # 44500 - 44000, seeded from the June fill-up outside this month's range


@pytest.mark.asyncio
async def test_build_expense_workbook_excludes_other_months_and_reps(fake_client):
    fake_client.store["commerce_expenses"] = [
        _claim(date="2026-06-10"),  # different month
        _claim(date="2026-07-10", paid_by="27829999999"),  # different rep
    ]
    result = await es_mod.build_expense_workbook(TID, "Richard", REP_PHONE, "2026-07-01", "2026-07-31")
    assert result is None


# ── is_due ────────────────────────────────────────────────────────────────────────

def test_is_due_true_on_configured_day():
    rep = _rep(expense_sheet_day_of_month=1)
    now = datetime(2026, 8, 1, 10, 0, tzinfo=timezone.utc)
    assert es_mod.is_due(rep, now) is True


def test_is_due_false_on_wrong_day():
    rep = _rep(expense_sheet_day_of_month=1)
    now = datetime(2026, 8, 2, 10, 0, tzinfo=timezone.utc)
    assert es_mod.is_due(rep, now) is False


def test_is_due_false_if_already_sent_this_month():
    now = datetime(2026, 8, 1, 10, 0, tzinfo=timezone.utc)
    rep = _rep(expense_sheet_day_of_month=1,
               expense_sheet_last_sent_at=(now - timedelta(hours=1)).isoformat())
    assert es_mod.is_due(rep, now) is False


def test_is_due_true_if_last_sent_a_different_month():
    now = datetime(2026, 8, 1, 10, 0, tzinfo=timezone.utc)
    rep = _rep(expense_sheet_day_of_month=1,
               expense_sheet_last_sent_at="2026-07-01T10:00:00+00:00")
    assert es_mod.is_due(rep, now) is True


# ── send_expense_sheet ────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_send_expense_sheet_success(fake_client, monkeypatch):
    fake_client.store["commerce_expenses"] = [_claim(date="2026-07-10")]
    monkeypatch.setattr(es_mod, "prior_month_range", lambda now=None: ("2026-07-01", "2026-07-31"))
    monkeypatch.setattr("vula.email_imap.credentials.get_email_creds",
                         lambda tid: {"email": "rep@gerflor.co.za"})

    sent = {}

    async def _fake_send(creds, to, subject, body, attachments=None):
        sent["to"] = to
        sent["attachments"] = attachments
        return {"sent": True}
    monkeypatch.setattr("vula.email_imap.service.send", _fake_send)

    ok = await es_mod.send_expense_sheet(TID, _rep())
    assert ok is True
    assert sent["to"] == "accounts@gerflor.co.za"
    assert len(sent["attachments"]) == 1
    assert sent["attachments"][0]["filename"].endswith(".xlsx")


@pytest.mark.asyncio
async def test_send_expense_sheet_no_recipient_configured(fake_client):
    ok = await es_mod.send_expense_sheet(TID, _rep(expense_sheet_recipient_email=None))
    assert ok is False


@pytest.mark.asyncio
async def test_send_expense_sheet_skips_when_no_claims(fake_client, monkeypatch):
    monkeypatch.setattr(es_mod, "prior_month_range", lambda now=None: ("2026-07-01", "2026-07-31"))
    ok = await es_mod.send_expense_sheet(TID, _rep())
    assert ok is False


@pytest.mark.asyncio
async def test_send_expense_sheet_no_connected_email_degrades(fake_client, monkeypatch):
    fake_client.store["commerce_expenses"] = [_claim(date="2026-07-10")]
    monkeypatch.setattr(es_mod, "prior_month_range", lambda now=None: ("2026-07-01", "2026-07-31"))
    monkeypatch.setattr("vula.email_imap.credentials.get_email_creds", lambda tid: None)
    # No IMAP mailbox — mail_router falls back to Microsoft Graph next; mock it as also not
    # connected so this test doesn't make a real Supabase call for vula_microsoft_accounts.
    monkeypatch.setattr("vula.microsoft.credentials.get_access_token", AsyncMock(return_value=None))
    ok = await es_mod.send_expense_sheet(TID, _rep())
    assert ok is False


# ── run_monthly_expense_sheets ────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_run_monthly_expense_sheets_skips_rep_with_no_recipient(fake_client):
    fake_client.store["vula_team_members"] = [_rep(expense_sheet_recipient_email=None)]
    sent = await es_mod.run_monthly_expense_sheets()
    assert sent == 0


@pytest.mark.asyncio
async def test_run_monthly_expense_sheets_skips_rep_not_due(fake_client, monkeypatch):
    monkeypatch.setattr(es_mod, "is_due", lambda rep, now: False)
    fake_client.store["vula_team_members"] = [_rep()]
    called = {"n": 0}

    async def _fake_send(*a, **kw):
        called["n"] += 1
        return True
    monkeypatch.setattr(es_mod, "send_expense_sheet", _fake_send)
    await es_mod.run_monthly_expense_sheets()
    assert called["n"] == 0


@pytest.mark.asyncio
async def test_run_monthly_expense_sheets_sends_and_stamps_when_due(fake_client, monkeypatch):
    monkeypatch.setattr(es_mod, "is_due", lambda rep, now: True)
    rep_row = _rep()
    fake_client.store["vula_team_members"] = [rep_row]

    async def _fake_send(tenant_id, rep):
        return True
    monkeypatch.setattr(es_mod, "send_expense_sheet", _fake_send)

    sent = await es_mod.run_monthly_expense_sheets()
    assert sent == 1
    updated_rep = [r for r in fake_client.store["vula_team_members"] if r["id"] == rep_row["id"]][0]
    assert updated_rep["expense_sheet_last_sent_at"] is not None


@pytest.mark.asyncio
async def test_run_monthly_expense_sheets_empty_month_still_stamps_due_cycle(fake_client, monkeypatch):
    monkeypatch.setattr(es_mod, "is_due", lambda rep, now: True)
    rep_row = _rep()
    fake_client.store["vula_team_members"] = [rep_row]

    async def _fake_send(tenant_id, rep):
        return False  # no claims this month
    monkeypatch.setattr(es_mod, "send_expense_sheet", _fake_send)

    sent = await es_mod.run_monthly_expense_sheets()
    assert sent == 0
    updated_rep = [r for r in fake_client.store["vula_team_members"] if r["id"] == rep_row["id"]][0]
    assert updated_rep["expense_sheet_last_sent_at"] is not None


@pytest.mark.asyncio
async def test_run_monthly_expense_sheets_one_rep_failure_does_not_block_another(fake_client, monkeypatch):
    monkeypatch.setattr(es_mod, "is_due", lambda rep, now: True)
    rep_a = _rep(id=str(uuid.uuid4()), whatsapp="27821111111")
    rep_b = _rep(id=str(uuid.uuid4()), whatsapp="27822222222")
    fake_client.store["vula_team_members"] = [rep_a, rep_b]

    async def _fake_send(tenant_id, rep):
        if rep["whatsapp"] == "27821111111":
            raise RuntimeError("boom")
        return True
    monkeypatch.setattr(es_mod, "send_expense_sheet", _fake_send)

    sent = await es_mod.run_monthly_expense_sheets()
    assert sent == 1
    updated_b = [r for r in fake_client.store["vula_team_members"] if r["id"] == rep_b["id"]][0]
    assert updated_b["expense_sheet_last_sent_at"] is not None
